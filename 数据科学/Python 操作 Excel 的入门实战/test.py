from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
#有个小院-兴趣编程

#读取工作表
wb=load_workbook('test.xlsx')
ws=wb.active

#读取某一数据
print(ws['B3'].value)

#读取一定范围的数据
for row in range(2,7):
    for col in range(1,4):
        #A2 B2 C2
        #A3 B3 C3 ...
        char=get_column_letter(col)
        print(ws[char+str(row) ].value)

#修改某一值
ws['A6'].value='小白'
wb.save('test.xlsx')

#创建工作表
wb.create_sheet('有个小院')

#多工作表是选择工作表
ws=wb['有个小院']

#合并格子
ws.merge_cells('D1:E1')

#取消合并
ws.unmerge_cells('D1:E1')
wb.save('test.xlsx')

#插入横排、竖排
ws.insert_rows(3)
ws.insert_cols(2)

#删除横排、竖排
ws.delete_rows(3)
ws.delete_cols(3)
wb.save('test.xlsx')

#移动资料
ws.move_range('A3:C4',rows=4,cols=5)
wb.save('test.xlsx')






