#有个小院-兴趣编程

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from  openpyxl.styles import Font

data=[ 
    {  
        'name':'张三',
        'heigth':181.6,
        'weight':74
    },
    {
        'name':'李四',
        'heigth':175.9,
        'weight':80
    },
    {
        'name':'赵一',
        'heigth':165.5,
        'weight':45
    },
    {
        'name':'牛二',
        'heigth':180.3,
        'weight':78
    },
    {
        'name':'孟五',
        'heigth':170.2,
        'weight':50
    }
]
#新建档案
wb=Workbook()
ws=wb.active
#添加标题
title=['姓名','身高','体重']
ws.append(title)

#插入数据
for person in data:
    #['张三','181.6','74']
    ws.append(list(person.values()))

#求平均值
for col in range(2,4):
   char=get_column_letter(col)
   #B7 C7
   ws[char+'7']=f'=AVERAGE({char+"2"}:{char+"6"})'

#修改标题字体
for col in range(1,4):
    char=get_column_letter(col)
    ws[char+'1'].font=Font(bold=True)

#保存
wb.save('new.xlsx')