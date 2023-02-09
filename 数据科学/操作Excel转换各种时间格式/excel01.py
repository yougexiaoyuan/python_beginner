'''
有个小院 - 兴趣编程
一分钟教你Python操作Excel转换各种时间格式
'''

#导入Excel模块
import openpyxl
#导入日期时间模块中的日期时间类
from datetime import datetime

#获取商品订单.xlsx文件对象
workbook=openpyxl.load_workbook('商品订单.xlsx')
#获取Sheet1表格对象
sheet=workbook['Sheet1']

#循环遍历Sheet1表格对象所有行
for rowNum in range(2,sheet.max_row+1):
    #获取当前行第2列单元格对象
    cell=sheet.cell(row=rowNum,column=2)
    #将单元格中的字符串值按照格式转换成日期时间对象
    time=datetime.strptime(cell.value,'%d/%m/%Y %I:%M:%S %p')
    #将新的格式日期写入当前行第3列单元格
    sheet.cell(row=rowNum, column=3).value=time
    #将中文格式日期写入当前行第4列单元格
    sheet.cell(row=rowNum, column=4).value=f'{time.year}年{time.month}月{time.day}日{time.hour}时{time.minute}分{time.second}秒'

#保存商品订单.xlsx文件对象
workbook.save('商品订单.xlsx')

